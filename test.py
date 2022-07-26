import pandas as pd
from openpyxl import load_workbook

# File location:
# excel_file_location = r'/Users/auarangu/Downloads/Wayne_CIQ_v24.xlsx'
excel_file_location = r'/Users/auarangu/Downloads/Wayne_CIQ_21_CP_DATA_CNSGW_v1.xlsx'

# Load the excel file to a wb
wb = load_workbook(filename=excel_file_location, data_only=True)

# Empty list to append the names of the sheets to be used
sheets_to_use = []

# Append the names to the empty list
[sheets_to_use.append(sheet) for sheet in wb.sheetnames if 'IP-Addresses' in sheet]


########################################################################################################################

def extract_data(data):
    rows_list = []

    for row in data:
        cols = []
        for col in row:
            cols.append(col.value)
        rows_list.append(cols)
    return rows_list


def evaluate_sheet(workbook, sheets_names, data_fringe):
    asd = []
    for i in range(1, len(sheets_names)):

        sheet = workbook[sheets_names[i]]
        data = sheet[data_fringe]
        rows = extract_data(data)
        df = pd.DataFrame(data=rows[1:], index=None, columns=rows[0])

        temp = df.iloc[:, [0, 2]]

        df1 = temp[temp['VlanX.K8.bond0'].notnull()]

        curr_rack = df1[df1['KVM Node Name'].notnull()]

        curr_rack = curr_rack[curr_rack["VlanX.K8.bond0"].str.contains("NA") == False]
        folder = "WAYNE/Rack " + str(i+2)

        curr_rack['folder'] = folder
        curr_rack['Protocol'] = "SSH2"
        curr_rack['username'] = "admin"
        curr_rack.loc[curr_rack['KVM Node Name'].str.contains("CM"), 'username'] = "cloud-user"
        curr_rack.loc[curr_rack['KVM Node Name'].str.contains("AI"), 'username'] = "cloud-user"
        curr_rack['emulation'] = 'XTerm'
        curr_rack.rename({'VlanX.K8.bond0': "hostname", "KVM Node Name": "session_name"}, axis=1, inplace=True)

        asd.append(curr_rack)

    return asd


extracted_df = []

# extracted_df = evaluate_sheet(wb, sheets_to_use, "B59:G154")


# 59B:154G
# Select the sheet to extract the data from.
sheet = wb[sheets_to_use[0]]

# Set the range of the table within the sheet from which data is going to be extracted
data = sheet['C35:H101']

# Empty list to store rows
rows_list = extract_data(data)

# Create the Pandas Data Frame to tabularize the data
df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])

temp = df.iloc[:, [0, 1, 3]]

df1 = temp[temp['Vlan101.K8.bond0'].notnull()]
df2 = df1[df1['Element ID'].notnull()]

# TO FUNCTION
rack_1 = df2[df2['K8 Node -Function'].str.contains('RACK1|Rack1|rack1')]
rack1 = rack_1.iloc[:, [0, 2]]
route_r1 = "WAYNE/Rack 1-CP"
rack1['folder'] = route_r1
rack1['Protocol'] = "SSH2"
rack1['username'] = "admin"
rack1.loc[rack1['Element ID'].str.contains("CM"), 'username'] = "cloud-user"
rack1.loc[rack1['Element ID'].str.contains("AI"), 'username'] = "cloud-user"
rack1['emulation'] = 'XTerm'
rack1.rename({'Vlan101.K8.bond0': "hostname", "Element ID": "session_name"}, axis=1, inplace=True)

########################################################################################################################


rack_2 = df2[df2['K8 Node -Function'].str.contains('RACK2|Rack2|rack2')]
rack2 = rack_2.iloc[:, [0, 2]]
route_r2 = "WAYNE/Rack 2-CP"
rack2['folder'] = route_r2
rack2['Protocol'] = "SSH2"
rack2['username'] = "admin"
rack2.loc[rack2['Element ID'].str.contains("CM"), 'username'] = "cloud-user"
rack2.loc[rack2['Element ID'].str.contains("AI"), 'username'] = "cloud-user"
rack2['emulation'] = 'XTerm'
rack2.rename({'Vlan101.K8.bond0': "hostname", "Element ID": "session_name"}, axis=1, inplace=True)

extracted_df.insert(0, rack1)
extracted_df.insert(1, rack2)


po = pd.concat(extracted_df)

print(po)

po.to_csv('extracted_info-CP.csv', index=False)